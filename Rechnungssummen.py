import sqlite3 #import der lib
from openpyxl import Workbook #import der lib

conn = sqlite3.connect("Chinook_Sqlite_AutoIncrementPKs.sqlite")
cur = conn.cursor()
# Ansprechen der lokalen sqlite Datenbank
year_ = ["2010", "2011", "2012", "2013"]
wb = Workbook()
ws1 = wb.active
ws1.title = "Rechnungssummen pro Jahr"
list_data = []
counter = 2
_ = ws1.cell(column=1, row=1, value="Jahr")
_ = ws1.cell(column=2, row=1, value="Name")
_ = ws1.cell(column=3, row=1, value="Rechnungssumme")

for item_year in year_:
    cur.execute(f"  SELECT\
                    Invoice.CustomerId, \
                    (SELECT LastName ||', ' || FirstName \
                        FROM Customer \
                        WHERE Customer.CustomerId = Invoice.CustomerId)\
                        AS Name,\
                        sum(Invoice.total) as Summe \
                    FROM \
                        Invoice\
                    WHERE Invoice.InvoiceDate like '{item_year}%'\
                    GROUP by \
                        Invoice.CustomerId;")
    list_data.append(cur.fetchall())
    for row in range(len(list_data)):
        for item in list_data[row]:
            a, b, c = item
            _ = ws1.cell(column=1, row=counter, value=item_year)
            _ = ws1.cell(column=2, row=counter, value=b)
            _ = ws1.cell(column=3, row=counter, value=c)
            counter += 1     
wb.save("Rech_sum.xlsx")
conn.close() 
