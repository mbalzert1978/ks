import sqlite3, pprint, argparse #import der lib
from openpyxl import Workbook #import der lib

parser = argparse.ArgumentParser()
parser.add_argument('--search', action='store',
                    dest='search_',
                    help='Store a table')
searching_ = parser.parse_args()


conn = sqlite3.connect("Chinook_Sqlite_AutoIncrementPKs.sqlite")
cur = conn.cursor()
# Ansprechen der lokalen sqlite Datenbank
cur.execute(f"select name as Songtitel,                                                 \
                (select Artist.Name from Artist                                         \
                    WHERE Track.GenreId = Artist.ArtistId) AS Kuenstler,                \
                (select MediaType.Name from MediaType                                   \
                    WHERE Track.GenreId = MediaType.MediaTypeId) AS Mediatypname,		\
                (select Genre.Name from Genre                                           \
                    WHERE Track.GenreId = Genre.GenreId) AS Genrename,                  \
                (select Album.Title from Album                                          \
                    WHERE Track.GenreId = Album.AlbumId) AS Albumname                   \
                from Track                                                              \
                where name like '%Rock%'                                \
                order by AlbumId                                                        \
                        ;")
#{searching_.search_}

list_data = cur.fetchall()
hilfs_var = 1
counter = 1
for item in list_data:
    a, b, c, d, e = item
    if hilfs_var == 1:
        wb = Workbook()
    if len(e) > 15:
        e = e[0:14]
        ws = wb.active
        ws.title = f"{e}"
    if hilfs_var == 1:
        hilfs_var = e
        _ = ws.cell(column=1, row=counter, value="Songtitle")
        _ = ws.cell(column=2, row=counter, value="Künstler")
        _ = ws.cell(column=3, row=counter, value="MediaTyp")
        _ = ws.cell(column=4, row=counter, value="Genre")
        counter += 1
        _ = ws.cell(column=1, row=counter, value=a)
        _ = ws.cell(column=2, row=counter, value=b)
        _ = ws.cell(column=3, row=counter, value=c)
        _ = ws.cell(column=4, row=counter, value=d)
    elif hilfs_var == e:
        counter += 1
        _ = ws.cell(column=1, row=counter, value=a)
        _ = ws.cell(column=2, row=counter, value=b)
        _ = ws.cell(column=3, row=counter, value=c)
        _ = ws.cell(column=4, row=counter, value=d)
    else:
        counter = 1
        wb.create_sheet(f"{e}")
        _ = ws.cell(column=1, row=counter, value="Songtitle")
        _ = ws.cell(column=2, row=counter, value="Künstler")
        _ = ws.cell(column=3, row=counter, value="MediaTyp")
        _ = ws.cell(column=4, row=counter, value="Genre")
        counter += 1
        _ = ws.cell(column=1, row=counter, value=a)
        _ = ws.cell(column=2, row=counter, value=b)
        _ = ws.cell(column=3, row=counter, value=c)
        _ = ws.cell(column=4, row=counter, value=d)
        counter = 1
        hilfs_var = e
wb.save(f"Rock.xlsx")
conn.close() 
